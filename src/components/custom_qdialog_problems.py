# -*- encoding: utf-8 -*-
'''
@File         :custom_qdialog_problems.py
@Time         :2025/06/12 13:55:36
@Author       :diamond_cz@163.com
@Version      :1.0
@Description  :自定义问题点记录对话框
'''

"""导入python内置模块"""
import os
import json

"""导入python第三方模块"""
import openpyxl
import win32com.client as win32
from PyQt5.QtCore import Qt
from PyQt5.QtWidgets import (QLabel, QComboBox, QLineEdit, QPushButton, QDialogButtonBox, 
                             QCheckBox, QVBoxLayout, QHBoxLayout, QDialog, QFileDialog)

"""导入项目自定义模块"""
from src.utils.stitchimg import stitch_images  # 导入拼接图片工具
from src.common.font_manager import SingleFontManager  # 导入字体管理器
from src.components.custom_qMbox_showinfo import show_message_box  # 导入消息框类

"""设置本项目的入口路径,全局变量BasePath"""
# 方法一：手动找寻上级目录，获取项目入口路径，支持单独运行该模块
if True:
    BasePath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# 方法二：直接读取主函数的路径，获取项目入口目录,只适用于hiviewer.py同级目录下的py文件调用
if False: # 暂时禁用，不支持单独运行该模块
    BasePath = os.path.dirname(os.path.abspath(sys.argv[0]))    


class ProblemsDialog(QDialog):
    """自定义对话框类, 用于输入信息"""
    def __init__(self, images_path_list, parent=None):
        super().__init__(parent)

        # 设置顶层窗口，方便调用类SubMainWindow中的函数与变量；根据传入的图片路径列表设置关联图片下拉框；
        self.parent_SubMainWindow = parent
        self.images_path_list = images_path_list

        # 初始化对话框UI
        self.init_ui()

        # 初始化下拉框combo_box0，设置关联项为图片名称
        self.init_combo_box()

        # 加载上次保存的设置
        self.load_settings()

        # 设置快捷键和槽函数
        self.set_shortcut()


    def init_ui(self):
        """初始化对话框UI"""

        # 设置窗口标题
        self.setWindowTitle("Camera Test 问题点记录")
        # 设置窗口大小
        self.setFixedSize(1200, 600)  # 设置对话框大小
        
        # 设置json保存路径
        self.save_path = os.path.join(BasePath, "cache", "test_settings.json")
        
        # 初始化字体管理器，标签组件使用;设置全局变量，定义项目基础路径
        self.font_manager_jetbrains_big = (self.parent_SubMainWindow.font_manager_j12 if self.parent_SubMainWindow and self.parent_SubMainWindow.font_manager_j12 
                                           else SingleFontManager.get_font(size=12, font_path=os.path.join(BasePath, "resource", "fonts", "JetBrainsMapleMono_Regular.ttf")))
        self.font_manager_jetbrains_small = (self.parent_SubMainWindow.font_manager_j10 if self.parent_SubMainWindow and self.parent_SubMainWindow.font_manager_j10 
                                             else SingleFontManager.get_font(size=10, font_path=os.path.join(BasePath, "resource", "fonts", "JetBrainsMapleMono_Regular.ttf")))
        
        # 创建主布局
        self.layout = QVBoxLayout(self)

        # 统一的下拉框高度和下拉框样式表
        combo_box_height = 35
        combobox_style = f"""
            QComboBox {{
                /* 下拉框本体样式 */
                min-height: 35px;
                font-family: "{self.font_manager_jetbrains_small.family()}";
                font-size: {self.font_manager_jetbrains_small.pointSize()}pt;
            }}
            
            QComboBox QAbstractItemView {{
                /* 下拉列表样式 */
                font-family: "{self.font_manager_jetbrains_small.family()}";
                font-size: {self.font_manager_jetbrains_small.pointSize()}pt;
            }}
            
            QComboBox QAbstractItemView::item {{
                /* 下拉项样式 */
                min-height: 35px;
                padding: 5px;
                font-family: "{self.font_manager_jetbrains_small.family()}";
                font-size: {self.font_manager_jetbrains_small.pointSize()}pt;
            }}

        """

        # 第零行：标签 + 下拉框 + 输入框
        layout_zero = QHBoxLayout()
        self.label0 = QLabel("关联图片项:", self)
        self.label0.setFont(self.font_manager_jetbrains_big)
        self.combo_box0 = QComboBox(self)
        self.combo_box0.setStyleSheet(combobox_style)
        self.text_input0 = QLineEdit(self)
        self.text_input0.setFont(self.font_manager_jetbrains_small)
        self.text_input0.setFixedHeight(combo_box_height)  # 设置输入框高度
        layout_zero.addWidget(self.label0)
        layout_zero.addWidget(self.combo_box0)
        layout_zero.addWidget(self.text_input0)
        # 设置比例
        layout_zero.setStretch(0, 1)  # label0 的比例
        layout_zero.setStretch(1, 4)  # combo_box0 的比例
        layout_zero.setStretch(2, 6)  # text_input0 的比例
        self.layout.addLayout(layout_zero)
        
        # 第一行：标签 + 输入框 + 加载按钮 + 写入按钮
        layout_one = QHBoxLayout()
        self.label1 = QLabel("问题点路径:", self)
        self.label1.setFont(self.font_manager_jetbrains_big)
        self.text_input1 = QLineEdit(self)
        self.text_input1.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.text_input1.setFont(self.font_manager_jetbrains_small)
        self.text_input1.setPlaceholderText("输入保存问题点的EXCEL路径...")  # 设置提示文本
        self.load_button = QPushButton("导入_Excel路径", self)
        self.load_button.setFont(self.font_manager_jetbrains_big)
        self.load_button.setFixedHeight(combo_box_height)  # 设置下拉框高度
        layout_one.addWidget(self.label1)
        layout_one.addWidget(self.text_input1)
        layout_one.addWidget(self.load_button)
        # 设置比例
        layout_one.setStretch(0, 1)  
        layout_one.setStretch(1, 10)
        layout_one.setStretch(2, 1)   
        self.layout.addLayout(layout_one)


        # 第二-1行：标签 + 输入框 + 汇总按钮 + 保存按钮
        layout_two_1 = QHBoxLayout()
        self.label2_1 = QLabel("拼接图路径:", self)
        self.label2_1.setFont(self.font_manager_jetbrains_big)
        self.text_input2_1 = QLineEdit(self)
        self.text_input2_1.setFixedHeight(combo_box_height)  
        self.text_input2_1.setFont(self.font_manager_jetbrains_small)
        self.text_input2_1.setPlaceholderText("输入或设置拼接图保存路径...")
        self.set_path_button_1 = QPushButton("设置_保存路径", self)
        self.set_path_button_1.setFont(self.font_manager_jetbrains_big)
        self.set_path_button_1.setFixedHeight(combo_box_height) 
        layout_two_1.addWidget(self.label2_1)
        layout_two_1.addWidget(self.text_input2_1)
        layout_two_1.addWidget(self.set_path_button_1)
        # 设置比例
        layout_two_1.setStretch(0, 1)   
        layout_two_1.setStretch(1, 10)
        layout_two_1.setStretch(2, 1)   
        self.layout.addLayout(layout_two_1)

        # 第二行：标签 + 输入框 + 汇总按钮 + 保存按钮
        layout_two = QHBoxLayout()
        self.label2 = QLabel("问题点汇总:", self)
        self.label2.setFont(self.font_manager_jetbrains_big)
        self.text_input2 = QLineEdit(self)
        self.text_input2.setFixedHeight(combo_box_height)  # 设置下拉框高度
        self.text_input2.setFont(self.font_manager_jetbrains_small)
        self.text_input2.setPlaceholderText("汇总显示各模块问题点...")
        self.save_button = QPushButton("保存_拼接图", self)
        self.save_button.setFont(self.font_manager_jetbrains_big)
        self.save_button.setFixedHeight(combo_box_height)
        layout_two.addWidget(self.label2)
        layout_two.addWidget(self.text_input2)
        layout_two.addWidget(self.save_button)
        # 设置比例
        layout_two.setStretch(0, 1)   # label2 的比例
        layout_two.setStretch(1, 10)  # combo_box2 的比例
        layout_two.setStretch(2, 1)   # save_button 的比例
        self.layout.addLayout(layout_two)


        # 第三行：复选框 + 输入框
        layout_three = QHBoxLayout()
        self.checkbox1 = QCheckBox("AE", self)
        self.checkbox1.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox1.setChecked(True)
        self.combo_box3 = QComboBox(self)
        self.combo_box3.setStyleSheet(combobox_style)
        self.combo_box3.setEditable(True)  
        layout_three.addWidget(self.checkbox1)
        layout_three.addWidget(self.combo_box3)
        # 设置比例
        layout_three.setStretch(0, 1)
        layout_three.setStretch(1, 10) 
        self.layout.addLayout(layout_three)

        # 第四行：复选框 + 输入框
        layout_four = QHBoxLayout()
        self.checkbox2 = QCheckBox("AWB", self)
        self.checkbox2.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox2.setChecked(True)
        self.combo_box4 = QComboBox(self)
        self.combo_box4.setStyleSheet(combobox_style)
        self.combo_box4.setEditable(True)  
        layout_four.addWidget(self.checkbox2)
        layout_four.addWidget(self.combo_box4)
        # 设置比例
        layout_four.setStretch(0, 1)
        layout_four.setStretch(1, 10)  
        self.layout.addLayout(layout_four)

        # 第五行：复选框 + 输入框
        layout_five = QHBoxLayout()
        self.checkbox3 = QCheckBox("AF", self)
        self.checkbox3.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox3.setChecked(True)
        self.combo_box5 = QComboBox(self)
        self.combo_box5.setStyleSheet(combobox_style)
        self.combo_box5.setEditable(True) 
        layout_five.addWidget(self.checkbox3)
        layout_five.addWidget(self.combo_box5)
        # 设置比例
        layout_five.setStretch(0, 1)
        layout_five.setStretch(1, 10)  
        self.layout.addLayout(layout_five)

        # 第六行：复选框 + 输入框
        layout_six = QHBoxLayout()
        self.checkbox4 = QCheckBox("ISP", self)
        self.checkbox4.setFont(self.font_manager_jetbrains_small)
        # 设置复选框的初始状态
        self.checkbox4.setChecked(True)
        self.combo_box6 = QComboBox(self)
        self.combo_box6.setStyleSheet(combobox_style)
        self.combo_box6.setEditable(True)  
        layout_six.addWidget(self.checkbox4)
        layout_six.addWidget(self.combo_box6)
        # 设置比例
        layout_six.setStretch(0, 1)
        layout_six.setStretch(1, 10)  
        self.layout.addLayout(layout_six)

        # 添加确认和取消按钮
        self.button_box = QDialogButtonBox(QDialogButtonBox.Ok | QDialogButtonBox.Cancel, self)
        self.button_box.setStyleSheet(combobox_style)
        self.layout.addWidget(self.button_box)


    def init_combo_box(self):
        """初始化下拉框combo_box0，设置关联项为图片名称"""
        # 设置是否在同一文件夹下 self.is_same_folder
        if self.images_path_list:
            self.is_same_folder = len(set([os.path.dirname(path) for path in self.images_path_list])) == 1
        
        # 先判断是否存在变量self.is_same_folder
        if hasattr(self, "is_same_folder"): 
            # 判断self.is_same_folder是否为True
            if self.is_same_folder:
                # 将 上一级文件夹名+图片名称 设置到 combo_box0 中
                image_names_all = [os.path.relpath(path, start=os.path.dirname(os.path.dirname(path))) for path in self.images_path_list]
                self.combo_box0.addItems(image_names_all)
            else:
                parent_folders = [os.path.basename(os.path.dirname(path)) for path in self.images_path_list]
                image_names = [os.path.basename(path) for path in self.images_path_list]
                # 将 上一级文件夹名 设置到 combo_box0 中
                self.combo_box0.addItems(parent_folders)
                # 设置字典将上一级文件夹名和图片名称对应起来
                self.parent_folder_dict = dict(zip(parent_folders, image_names))
        else:
            print(f"ProblemsDialog: 传入的图片路径列表为空, 请检查传入的图片路径列表")


    def save_settings(self):
        """保存当前设置"""
        settings = {
            "关联图片项": self.combo_box0.currentText(),
            "问题点路径": self.text_input1.text(),
            "拼接图路径": self.text_input2_1.text(),
            "AE": [self.combo_box3.itemText(i) for i in range(self.combo_box3.count())],
            "AWB": [self.combo_box4.itemText(i) for i in range(self.combo_box4.count())],
            "AF": [self.combo_box5.itemText(i) for i in range(self.combo_box5.count())],
            "ISP": [self.combo_box6.itemText(i) for i in range(self.combo_box6.count())],
        }
        with open(self.save_path, 'w', encoding='utf-8') as f:
            json.dump(settings, f, ensure_ascii=False, indent=4)

    def load_settings(self):
        """加载上次保存的设置"""
        try:
            with open(self.save_path, 'r', encoding='utf-8') as f:
                settings = json.load(f)
                combo_box0_text = settings.get("关联图片项", "")
                combo_box0_items = [self.combo_box0.itemText(i) for i in range(self.combo_box0.count())]
                if combo_box0_text and combo_box0_text in combo_box0_items:
                    # 如果 combo_box0 的当前文本不为空，则更新 text_input0 的文本
                    self.combo_box0.setCurrentText(combo_box0_text)
                    self.update_input0_text()
                
                self.text_input1.setText(settings.get("问题点路径", ""))
                self.text_input2_1.setText(settings.get("拼接图路径", ""))

                # 恢复AE、AWB、AF、ISP的选项
                for item in settings.get("AE", []):
                    self.combo_box3.addItem(item)
                for item in settings.get("AWB", []):
                    self.combo_box4.addItem(item)
                for item in settings.get("AF", []):
                    self.combo_box5.addItem(item)
                for item in settings.get("ISP", []):
                    self.combo_box6.addItem(item)
        except FileNotFoundError:
            print("未找到设置文件，使用默认值。")
        except json.JSONDecodeError:
            print("设置文件格式错误，使用默认值。")


    def set_shortcut(self):
        """设置快捷键和槽函数"""
        self.button_box.accepted.connect(self.accept)               # 连接确认按钮信号
        self.button_box.rejected.connect(self.reject)               # 连接取消按钮信号
        self.combo_box0.activated.connect(self.update_input0_text)  # 连接 combo_box0 的信号
        self.load_button.clicked.connect(self.load_data)            # 连接加载按钮信号
        self.set_path_button_1.clicked.connect(self.set_path_data)  # 连接设置路径按钮信号
        self.save_button.clicked.connect(self.save_stitch_images)   # 连接保存拼接图按钮信号
        self.finished.connect(self.save_settings)                   # 连接关闭信号

        # 输入框内容更新信号
        self.combo_box3.currentTextChanged.connect(self.refresh_data)
        self.combo_box4.currentTextChanged.connect(self.refresh_data)
        self.combo_box5.currentTextChanged.connect(self.refresh_data)
        self.combo_box6.currentTextChanged.connect(self.refresh_data)


    def update_input0_text(self):   
        """combo_box0 的信号"""
        # 获取 combo_box0 的当前文本
        current_text = self.combo_box0.currentText()
        # 判断当前文本是否为空
        if current_text and hasattr(self, "is_same_folder"):
            if self.is_same_folder: # 同一文件夹下 current_text 为 上一级文件夹名+图片名称
                # 将解析后的图片名称直接设置到 text_input0 中
                self.text_input0.setText(current_text.split("\\")[-1])
            else: # 不同文件夹下 current_text 为 上一级文件夹名
                # 将对应上一级文件夹内的图片名称设置到 text_input0 中
                if hasattr(self, "parent_folder_dict"):
                    self.text_input0.setText(self.parent_folder_dict[current_text])
                else:
                    print(f"ProblemsDialog: 父文件夹字典self.parent_folder_dict不存在, 请检查self.parent_folder_dict")


    def refresh_data(self):
        """汇总数据"""
        # 汇总3A+ISP问题点
        items = []
        if self.checkbox1.isChecked() and self.combo_box3.currentText():
            items.append(f"AE{self.combo_box3.currentText()}")
        if self.checkbox2.isChecked() and self.combo_box4.currentText():
            items.append(f"AWB{self.combo_box4.currentText()}")
        if self.checkbox3.isChecked() and self.combo_box5.currentText():
            items.append(f"AF{self.combo_box5.currentText()}")
        if self.checkbox4.isChecked() and self.combo_box6.currentText():
            items.append(f"ISP{self.combo_box6.currentText()}")
        # 拼接内容并设置到 combo_box2
        self.text_input2.clear()  # 清空现有内容
        _problems = f"{'_'.join(items)}.png"
        self.text_input2.setText(_problems)  

    def close_excel(self):
        """
        该函数主要是实现了一个 强制关闭一个EXCEL表格 的功能.
        """
        try:
            # 获取 Excel 应用程序的实例
            excel_app = win32.gencache.EnsureDispatch('Excel.Application')
            
            # 检查是否有打开的工作簿
            if excel_app.Workbooks.Count > 0:
                # 强制关闭所有工作簿，不保存更改
                for workbook in excel_app.Workbooks:
                    workbook.Close(SaveChanges=False)
            
            # 退出 Excel 应用程序
            excel_app.Quit()
            print("Excel 已成功关闭。")
            
        except Exception as e:
            print(f"关闭 Excel 时发生错误: {e}")


    def save_stitch_images(self):
        """读取问题点汇总描述,拼接图片,以上述描述为文件名,保存拼接图到指定路径"""
        # 获取拼接图路径和问题点汇总描述
        problem_description = self.text_input2.text()
        save_path = self.text_input2_1.text()
        # 判断save_path和problem_description是否为空
        if not save_path or not problem_description:
            show_message_box("请输入拼接图路径和问题点汇总描述", "提示", 1500)
            return
        # 判断save_path是否存在,不存在则创建save_path
        if not os.path.exists(save_path):
            os.makedirs(save_path)
        # 调用拼接图片函数
        image_paths = os.path.join(save_path, problem_description)
        # 设置字体大小为12，与界面字体保持一致
        sucess = stitch_images(self.images_path_list, image_paths, self.font_manager_jetbrains_big)
        show_message_box("拼接图片成功!", "提示", 1500) if sucess else show_message_box("拼接图片失败!", "提示", 1500)


    
    def write_data_backup(self):
        """写入数据,不支持实时写入,初始版本"""  
        excel_path = self.text_input1.text()  # 获取Excel文件路径
        search_value = self.text_input0.text()  # 获取要查找的值

        # 判断excel_path是否为空
        if not excel_path:
            print("Excel文件路径为空，请输入Excel文件路径。")
            return

        # 判断search_value是否为空
        if not search_value:
            print("要查找的值为空，请输入要查找的值。")
            return

        #尝试打开文件判断是否被占用,若被占用了则强制关闭excel
        try:
            with open(excel_path, 'a') as f:
                pass
        except PermissionError:
            print(f"Error: Excel file is locked or in use: {excel_path}")
            self.close_excel()  

        try:
            # 对excel表格进行读写操作
            workbook = openpyxl.load_workbook(excel_path) 
            sheet = workbook['Sheet1'] # workbook.active

            # 获取第一行的值
            first_row_values = [cell.value for cell in sheet[1]]  # sheet[1]表示第一行
            target_column_ae = first_row_values.index("AE") + 1
            problem_ae = self.combo_box3.currentText()
            target_column_awb = first_row_values.index("AWB") + 1
            problem_awb = self.combo_box4.currentText()
            target_column_af = first_row_values.index("AF") + 1
            problem_af = self.combo_box5.currentText()
            target_column_isp = first_row_values.index("ISP") + 1
            problem_isp = self.combo_box6.currentText()

            # 获取第一列的值
            first_column_values = [cell.value for cell in sheet['A']]  # 'A'表示第一列

            # search_value匹配，获取目标行索引
            target_row = 0
            for index, value in enumerate(first_column_values):
                if value == search_value:
                    target_row  = index + 1
            if not target_row: # 如果找不到对应的图片行，目标索引就是没有数据的最后一行
                target_row = len(sheet['A']) + 1
                # 写入对应的图片名称
                sheet.cell(target_row, 1).value = search_value
                
            # 向表格写入数据
            sheet.cell(target_row, target_column_ae).value = problem_ae
            sheet.cell(target_row, target_column_awb).value = problem_awb
            sheet.cell(target_row, target_column_af).value = problem_af
            sheet.cell(target_row, target_column_isp).value = problem_isp
        
            # 保存修改后的文件
            workbook.save(excel_path)
            print("数据已成功写入Excel文件。")

        except Exception as e:
            print(f"写入数据时发生错误: {e}")


    def write_data(self):
        """写入数据，支持实时写入"""  
        excel_path = self.text_input1.text()  # 获取Excel文件路径
        search_value = self.text_input0.text()  # 获取要查找的值

        # 判断excel_path是否为空
        if not os.path.exists(excel_path):
            show_message_box("Excel文件路径为空或不存在, 请加载或手动输入Excel文件路径", "提示", 1500)
            return

        if not search_value:
            show_message_box("没有要关联的图片项, 请点击下拉框加载", "提示", 1500)
            return

        try:
            # 启动 Excel 应用程序
            excel_app = win32.gencache.EnsureDispatch('Excel.Application')
            excel_app.Visible = True  # 设置为可见

            # 打开工作簿
            workbook = excel_app.Workbooks.Open(excel_path)
            # 选择第一个工作表
            sheet = workbook.Worksheets(1)  # 索引从 1 开始

            # 查找目标行
            target_row = 0
            for row in range(1, sheet.UsedRange.Rows.Count + 1):
                if sheet.Cells(row, 1).Value == search_value:  # 假设在第一列查找
                    target_row = row
                    break

            if target_row == 0:  # 如果找不到对应的行
                target_row = sheet.UsedRange.Rows.Count + 1      # 写入到最后一行
                sheet.Cells(target_row, 1).Value = search_value  # 写入文件名


            # 找到目标行后；按列，向表格写入数据
            for col in range(1, sheet.UsedRange.Columns.Count + 1):
                if sheet.Cells(1, col).Value  == "AE":
                    sheet.Cells(target_row, col).Value = self.combo_box3.currentText()
                if sheet.Cells(1, col).Value  == "AWB":
                    sheet.Cells(target_row, col).Value = self.combo_box4.currentText()
                if sheet.Cells(1, col).Value  == "AF":
                    sheet.Cells(target_row, col).Value = self.combo_box5.currentText()
                if sheet.Cells(1, col).Value  == "ISP":
                    sheet.Cells(target_row, col).Value = self.combo_box6.currentText()     

            # 保存工作簿
            workbook.Save()

            print("数据已成功写入Excel文件。")

        except Exception as e:
            print(f"写入数据时发生错误: {e}")
        finally:
            pass
            # 关闭工作簿（如果需要）
            # workbook.Close(SaveChanges=True)  # 如果需要保存更改
            # excel_app.Quit()  # 关闭 Excel 应用程序


    def load_data(self):
        """加载EXCEL表格"""
        try:
            options = QFileDialog.Options()
            _dir = next((os.path.dirname(path) for path in self.images_path_list if self.text_input0.text() in path), "")
            file_path, _ = QFileDialog.getOpenFileName(self, "选择EXCEL文件", _dir, "Excel Files (*.xls *.xlsx);;All Files (*)", options=options)
            if file_path:
                self.text_input1.setText(file_path)  # 显示选定的文件路径
        except Exception as e:
            print(f"加载EXCEL表格时发生错误: {e}")


    def set_path_data(self):
        """设置拼接图保存路径"""
        try:
            options = QFileDialog.Options()
            # 获取第一个图片的父文件夹作为默认路径
            _dir = next((os.path.dirname(path) for path in self.images_path_list if self.text_input0.text() in path), "")
            folder_path = QFileDialog.getExistingDirectory(self, "选择图片文件夹", os.path.dirname(_dir), options=options)  # 获取文件夹路径
            if folder_path:
                self.text_input2_1.setText(folder_path)  # 显示选定的文件路径
        except Exception as e:
            print(f"设置拼接图保存路径时发生错误: {e}") 

    def get_data(self):
        """获取用户输入的数据"""
        return {
            "关联图片项": self.combo_box0.currentText() + "/" + self.text_input0.text(),
            "问题点": self.text_input2.text(),
            "AE": self.combo_box3.currentText(),
            "AWB": self.combo_box4.currentText(),
            "AF": self.combo_box5.currentText(),
            "ISP": self.combo_box6.currentText(),
        }


    def keyPressEvent(self, event):
        """重写键盘按下事件，防止在输入框或下拉框中按下回车时关闭对话框"""
        if event.key() == Qt.Key_Return or event.key() == Qt.Key_Enter:
            # 如果当前焦点在输入框或下拉框中，阻止默认行为
            if self.focusWidget() in [self.text_input0, self.text_input2, self.combo_box0, self.combo_box3, self.combo_box4, self.combo_box5, self.combo_box6]:
                event.ignore()  # 忽略事件
            # 如果当前焦点在问题点输入框中，打开Excel文件
            elif self.focusWidget() == self.text_input1:
                excel_path = self.text_input1.text()
                if os.path.exists(excel_path):
                    os.startfile(excel_path)  # 在Windows上打开文件
                else:
                    print("指定的Excel文件不存在。")
            else:
                super().keyPressEvent(event)  # 处理其他情况
        else:
            super().keyPressEvent(event)  # 处理其他按键事件