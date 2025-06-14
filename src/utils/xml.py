# -*- coding: utf-8 -*-
import os
import sys
import xml.etree.ElementTree as ET
from pathlib import Path
from openpyxl import Workbook

"""设置本项目的入口路径,全局变量BasePath"""
# 方法一：手动找寻上级目录，获取项目入口路径，支持单独运行该模块
if True:
    BasePath = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
# 方法二：直接读取主函数的路径，获取项目入口目录,只适用于hiviewer.py同级目录下的py文件调用
if False: # 暂时禁用，不支持单独运行该模块
    BasePath = os.path.dirname(os.path.abspath(sys.argv[0]))  


def save_excel_data(images_path):
    """将从XML文件中提取的数据保存到Excel表格中"""

    # 配置保存的Excel文件路径
    excel_path = os.path.join(images_path, "extracted_data.xlsx")
    if os.path.exists(excel_path):
        # 若存在excel文件则不需要保存新的excel文件
        return

    # 初始化二维列表，第一行的表头数据
    get_excel_list = [
        [
            "文件名",
            "Lux",
            "DRCgain",
            "Safe_gain",
            "Short_gain",
            "Long_gain",
            "CCT",
            "R_gain",
            "B_gain",
            "Awb_sa",
            "Triangle_index",
            "AE",   # AE 问题点
            "AWB",  # AWB 问题点    
            "ISP",  # ISP 问题点
            "AF",   # AF 问题点
        ]
    ]

    # 遍历文件夹，列出所有满足条件的xml文件
    xml_files = [f for f in os.listdir(images_path) if f.endswith('_new.xml')]

    # 遍历xml文件
    for xml_file in xml_files:
        xml_file = str(Path(images_path) / xml_file)
        if os.path.exists(xml_file):
            # 使用函数load_xml_data加载并解析xml文件
            result_list = load_xml_data(xml_file)
            if result_list:
                get_excel_list.append(result_list)

    # 创建一个新的工作簿
    wb = Workbook()
    ws = wb.active

    # 设置边框样式
    from openpyxl.styles import Border, Side
    thin = Side(style='thin')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    # 写入数据并设置列宽和边框
    for row in get_excel_list:
        ws.append(row)
        for cell in row:
            # 设置边框
            ws.cell(row=ws.max_row, column=row.index(cell) + 1).border = border

    # 自适应列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter  # 获取列字母
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)  # 加2以增加一些额外的空间
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(excel_path)  # 保存为Excel文件
    print(f"数据已保存到 {excel_path}")


def load_xml_data(xml_path):
    """加载XML文件并提取Lux值和DRCgain值等EXIF信息"""
    try:
        # 加载xml文件
        tree = ET.parse(xml_path)
        root = tree.getroot()

        # 提取值并转换为列表
        result_list = [
            # 文件名
            str(os.path.basename(xml_path).split('_new.xml')[0]+".jpg"),
            # Lux
            float(root.find('lux_index').text) if root.find('lux_index') is not None else None,
            # DRCgain
            root.find('DRCgain').text if root.find('DRCgain') is not None else None,
            # Safe_gain
            float(root.find('safe_gain').text) if root.find('safe_gain') is not None else None,
            # Short_gain
            float(root.find('short_gain').text) if root.find('short_gain') is not None else None,
            # Long_gain
            float(root.find('long_gain').text) if root.find('long_gain') is not None else None,
            # CCT
            float(root.find('CCT').text) if root.find('CCT') is not None else None,
            # R_gain
            float(root.find('r_gain').text) if root.find('r_gain') is not None else None,
            # B_gain
            float(root.find('b_gain').text) if root.find('b_gain') is not None else None,
            # Awb_sa
            root.find('awb_sa').text if root.find('awb_sa') is not None else None,
            # Triangle_index
            float(root.find('triangle_index').text) if root.find('triangle_index') is not None else None,
        ]

        return result_list
        
    except Exception as e:
        print(f"解析XML失败{xml_path}:\n {e}")
        return None



